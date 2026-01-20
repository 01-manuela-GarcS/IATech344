# processor.py
# Lógica de negocio: operaciones sobre Excel

from openpyxl import load_workbook


def ejecutar_accion(instruccion, archivo_excel):
    # Puede ser    archivo_excel   o   path
    # Abre el archivo de ejemplo
    wb = load_workbook(archivo_excel)
    # podemos seleccionar el archvio que queremos reemplazar, con contrl h, podemos ver todas las concidencias con ese miemso nombre cambiarlas todas sin tener que ir una por una
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(archivo_excel)
 