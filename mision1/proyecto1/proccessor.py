import re
from openpyxl import load_workbook
# ====================
#  FUNCION clean_id
# Eliminar caracteres no númericos de un documento
# "cc75.888.56" = "7588856"
# ====================

def clean_id(value):
    # Elimina caracteres no númericos de un documento
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))
    
# Cunaod llamemos a la función no tiene ningun vlaor va a devolver vacio. Toma todo lo que no e sun digito y que lo quite, el str string. con '\D' toma numeros.
#El archivo processor.py debe estar acompañado de el archivo test_pro (test unitarios), debe tenr el noimbre test y el nombre al cual le estamos aplicando las pruebas. Es importante entregarlo para ver las pruebas que se realizaron

# ====================
#  FUNCION marge_name
# Une nombre y apellido en un solo campo
# ====================

def merge_name(name, lastname):
    if name is None:
        name =""
    if lastname is None:
        lastname =""
    return f"{name} {lastname}".strip()



def process_excel(path): #path es una ruta (ejem: Escritorio\IA_TalentoTech_344), en este caso es la ruta del excel que vamos a trabajar
    #Acceso a la hoja llamada "datos"
    wb= load_workbook(path)
    ws = wb["Datos"] #Esto representa una hoja
    #con +1 indicamos que no deseamos seleccionar la primera fila, es decir, que no tome los datos de la primera fila
    for row in range(2,ws.max_row+1):
        ws[f"D{row}"] =clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
        )
    # Guardar los cambios en el mismo archivo (path)
    wb.save(path)

def proccess_excel_safe(path):
    #Es lo que se va a ejecutar correctamente
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente."
    except PermissionError:
        return(
            False,
            "El archivo Excel está abierto.\n"
            # La \n indica el texto seguira en la siguiente linea. Este error sale si el archivo que se busca "ejecutar" esta abierto
            "Por favor, ciérrelo o intente nuevamente."
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        #Las que se pueden generar y que no conocemos. La identificamos con e
        return False, f"Error inesperado: {str(e)}"